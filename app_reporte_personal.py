from __future__ import annotations
import streamlit as st
import html
import io
import socket
import urllib.parse
from collections import Counter, defaultdict
from datetime import datetime
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
import openpyxl

# --- FUNCIÓN DE CONTROL DE ACCESO ---
def verificar_contrasena():
    if "autenticado" not in st.session_state:
        st.session_state["autenticado"] = False

    if st.session_state["autenticado"]:
        return True

    # Interfaz de Login elegante antes de cargar el reporte
    st.title("🔒 Acceso Restringido")
    st.subheader("Informe de Dotación Histórica")
    
    # Imagen de presentación (puedes cambiar este link por el tuyo)
    st.image("https://cdn-icons-png.flaticon.com/512/3064/3064155.png", width=120)
    
    clave_usuario = st.text_input("Introduce la contraseña para ver el reporte dinámico:", type="password")
    
    # Pon aquí la clave que tú quieras compartir
    if st.button("Ingresar al Reporte"):
        if clave_usuario == "MiPassword123":
            st.session_state["autenticado"] = True
            st.rerun()
        else:
            st.error("❌ Contraseña incorrecta.")
            
    return False

# --- COMPROBACIÓN DE SEGURIDAD ---
if verificar_contrasena():
    
    # -------------------------------------------------------------
    # TODO TU CÓDIGO ORIGINAL DEL REPORTE EMPIEZA AQUÍ:
    # -------------------------------------------------------------
    st.success("🔓 Acceso Concedido")
    st.title("📊 Reporte Dinámico de Dotación")
    st.write("El servidor ha ejecutado tus librerías y cargado los datos correctamente.")

import html
import io
import socket
import urllib.parse
from collections import Counter, defaultdict
from datetime import datetime
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path

import openpyxl


MOD_UBICACIONES = [
    "ARMADO 1 / PROD. SERIGRAFIA",
    "ARMADO 1/PROD.",
    "ARMADO 2/PROD.",
    "ARMADO 3 / PROD. SERIGRAFIA",
    "ARMADO 3/PROD.",
    "ARMADO 4/PROD.",
    "ARMADO 1/TRASLADO",
    "ARMADO 2/TERMO",
    "ARMADO 3/TRASLADO",
    "ARMADO 4/TRASLADO",
    "INYECCION DIURNA/PROD.",
    "INYECCION NOCTURNA/PROD.",
    "INYECCION TARDE/PROD.",
    "MECANIZADO DIURNO/PROD.",
    "MECANIZADO TARDE/PROD.",
]

STATE = {
    "filename": None,
    "rows": [],
}

TODAY = datetime(2026, 6, 5)


def esc(value):
    return html.escape("" if value is None else str(value))


def years_between(start):
    if not isinstance(start, datetime):
        return None
    return round((TODAY - start).days / 365.25, 1)


def parse_workbook(data: bytes, filename: str):
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[5]]
    rows = []
    for raw in ws.iter_rows(min_row=6, values_only=True):
        if not raw or raw[0] is None:
            continue
        item = dict(zip(headers, raw))
        item["Antiguedad"] = years_between(item.get("F. de Ingreso"))
        item["Es_MDO"] = item.get("Ubicación") in MOD_UBICACIONES
        item["Grupo_MDO"] = (
            "ARM" if "ARMADO" in str(item.get("Ubicación"))
            else "INY" if "INYECCION" in str(item.get("Ubicación"))
            else "MEC" if "MECANIZADO" in str(item.get("Ubicación"))
            else ""
        )
        rows.append(item)
    STATE["filename"] = filename
    STATE["rows"] = rows


def avg(values):
    nums = [v for v in values if isinstance(v, (int, float))]
    return round(sum(nums) / len(nums), 1) if nums else ""


def bar_table(title, rows, total, base_url, param):
    body = []
    max_count = max([count for _, count in rows], default=1)
    for label, count in rows:
        pct = count / total if total else 0
        width = int((count / max_count) * 100) if max_count else 0
        href = f"{base_url}?{urllib.parse.urlencode({param: label})}"
        body.append(
            f"<tr><td><a href='{href}'>{esc(label)}</a></td>"
            f"<td class='num'><a href='{href}'>{count}</a></td>"
            f"<td class='pct'>{pct:.1%}</td>"
            f"<td><div class='bar'><span style='width:{width}%'></span></div></td></tr>"
        )
    return (
        f"<section class='panel'><h2>{esc(title)}</h2>"
        "<table><thead><tr><th>Concepto</th><th>Cantidad</th><th>%</th><th>Visual</th></tr></thead>"
        f"<tbody>{''.join(body)}</tbody></table></section>"
    )


def people_table(rows):
    headers = [
        "Legajo",
        "Apellido y Nombre",
        "F. de Ingreso",
        "Sexo",
        "Tipo Personal",
        "Modalidad de Contratación",
        "Antiguedad",
        "Centro de Costo",
        "Ubicación",
    ]
    body = []
    for r in rows:
        cells = []
        for h in headers:
            value = r.get(h)
            if h == "F. de Ingreso" and isinstance(value, datetime):
                value = value.strftime("%d/%m/%Y")
            cells.append(f"<td>{esc(value)}</td>")
        body.append(f"<tr>{''.join(cells)}</tr>")
    return (
        "<table class='people'><thead><tr>"
        + "".join(f"<th>{esc(h)}</th>" for h in headers)
        + "</tr></thead><tbody>"
        + "".join(body)
        + "</tbody></table>"
    )


def layout(content, title="Reporte de Personal"):
    loaded = STATE["filename"]
    nav = ""
    if loaded:
        nav = (
            "<nav>"
            "<a href='/'>Dashboard</a>"
            "<a href='/mdo'>MDO</a>"
            "<a href='/general'>General</a>"
            "<a href='/antiguedad'>Antiguedad</a>"
            "</nav>"
        )
    return f"""<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{esc(title)}</title>
  <style>
    :root {{
      --bg:#e9f0f5; --ink:#172033; --muted:#667085; --line:#cbd5e1;
      --blue:#1f4e78; --green:#2f855a; --orange:#b7791f; --panel:#ffffff;
    }}
    * {{ box-sizing:border-box; }}
    body {{ margin:0; font-family:Segoe UI, Arial, sans-serif; background:var(--bg); color:var(--ink); }}
    header {{ background:var(--blue); color:white; padding:16px 24px; display:flex; align-items:center; gap:24px; }}
    header h1 {{ margin:0; font-size:20px; }}
    nav {{ display:flex; gap:8px; }}
    nav a {{ color:white; text-decoration:none; padding:7px 10px; border:1px solid rgba(255,255,255,.35); border-radius:6px; }}
    main {{ padding:20px 24px 36px; }}
    .upload {{ max-width:720px; margin:40px auto; background:var(--panel); padding:24px; border:1px solid var(--line); border-radius:8px; }}
    .grid {{ display:grid; grid-template-columns:repeat(12,1fr); gap:16px; align-items:start; }}
    .panel {{ background:var(--panel); border:1px solid var(--line); border-radius:8px; padding:14px; grid-column:span 6; }}
    .panel.full {{ grid-column:1 / -1; }}
    .kpi {{ background:var(--panel); border:1px solid var(--line); border-radius:8px; padding:14px; grid-column:span 3; }}
    .kpi b {{ display:block; font-size:24px; margin-top:6px; }}
    .kpi span {{ color:var(--muted); font-size:12px; font-weight:700; text-transform:uppercase; }}
    h2 {{ margin:0 0 10px; font-size:15px; color:var(--blue); }}
    table {{ width:100%; border-collapse:collapse; font-size:13px; background:white; }}
    th, td {{ padding:7px 8px; border-bottom:1px solid #e5e7eb; vertical-align:middle; }}
    th {{ background:#d9eaf7; color:#111827; position:sticky; top:0; z-index:1; }}
    a {{ color:#0954a5; font-weight:600; text-decoration:none; }}
    .num, .pct {{ text-align:right; white-space:nowrap; }}
    .bar {{ height:10px; background:#edf2f7; border-radius:999px; overflow:hidden; min-width:120px; }}
    .bar span {{ display:block; height:100%; background:#4472c4; }}
    .toolbar {{ display:flex; justify-content:space-between; align-items:center; gap:12px; margin-bottom:12px; }}
    .search {{ width:320px; max-width:100%; padding:8px 10px; border:1px solid var(--line); border-radius:6px; }}
    .people-wrapper {{ overflow:auto; max-height:72vh; border:1px solid var(--line); border-radius:8px; }}
    .people th {{ white-space:nowrap; }}
    .note {{ color:var(--muted); }}
    @media (max-width:900px) {{ .panel,.kpi {{ grid-column:1 / -1; }} header {{ display:block; }} nav {{ margin-top:12px; flex-wrap:wrap; }} }}
  </style>
</head>
<body>
  <header><h1>Reporte de Personal</h1>{nav}</header>
  <main>{content}</main>
  <script>
    const search = document.querySelector('[data-search]');
    if (search) {{
      search.addEventListener('input', () => {{
        const term = search.value.toLowerCase();
        document.querySelectorAll('tbody tr').forEach(tr => {{
          tr.style.display = tr.innerText.toLowerCase().includes(term) ? '' : 'none';
        }});
      }});
    }}
  </script>
</body>
</html>"""


def upload_page():
    return layout("""
    <section class="upload">
      <h2>Cargar input</h2>
      <p class="note">Seleccioná el Excel de legajos. El procesamiento se hace localmente en tu PC.</p>
      <form method="post" action="/upload" enctype="multipart/form-data">
        <input type="file" name="file" accept=".xlsx" required>
        <button type="submit">Procesar</button>
      </form>
    </section>
    """)


def dashboard():
    rows = STATE["rows"]
    if not rows:
        return upload_page()
    total = len(rows)
    mdo = [r for r in rows if r["Es_MDO"]]
    sexo = Counter(r["Sexo"] for r in rows)
    contrato = Counter(r["Modalidad de Contratación"] for r in rows)
    mdo_group = Counter(r["Grupo_MDO"] for r in mdo)
    content = [
        "<div class='grid'>",
        f"<div class='kpi'><span>Total dotacion</span><b>{total}</b></div>",
        f"<div class='kpi'><span>MDO</span><b><a href='/mdo'>{len(mdo)}</a></b></div>",
        f"<div class='kpi'><span>Antiguedad prom.</span><b>{avg([r['Antiguedad'] for r in rows])}</b></div>",
        f"<div class='kpi'><span>Antiguedad MDO</span><b>{avg([r['Antiguedad'] for r in mdo])}</b></div>",
        bar_table("MDO por grupo", mdo_group.most_common(), len(mdo), "/mdo", "grupo"),
        bar_table("Sexo", sexo.most_common(), total, "/detalle", "sexo"),
        bar_table("Modalidad de contratacion", contrato.most_common(), total, "/detalle", "contrato"),
        "</div>",
    ]
    return layout("".join(content))


def general_page():
    rows = STATE["rows"]
    if not rows:
        return upload_page()
    ubic = Counter(r["Ubicación"] for r in rows)
    cc = Counter(r["Centro de Costo"] for r in rows)
    return layout(
        "<div class='grid'>"
        + bar_table("Total por ubicacion", ubic.most_common(), len(rows), "/detalle", "ubicacion")
        + bar_table("Total por centro de costo", cc.most_common(), len(rows), "/detalle", "centro")
        + "</div>",
        "General",
    )


def mdo_page(query):
    rows = [r for r in STATE["rows"] if r["Es_MDO"]]
    if not rows:
        return upload_page()
    grupo = query.get("grupo", [""])[0]
    ubicacion = query.get("ubicacion", [""])[0]
    filtered = rows
    title = "MDO"
    if grupo:
        filtered = [r for r in filtered if r["Grupo_MDO"] == grupo]
        title = f"MDO | {grupo}"
    if ubicacion:
        filtered = [r for r in filtered if r["Ubicación"] == ubicacion]
        title = f"MDO | {ubicacion}"
    ubic = Counter(r["Ubicación"] for r in rows)
    content = (
        "<div class='grid'>"
        + bar_table("Ubicaciones MDO", ubic.most_common(), len(rows), "/mdo", "ubicacion")
        + f"<section class='panel full'><div class='toolbar'><h2>{esc(title)} ({len(filtered)})</h2>"
        + "<input class='search' data-search placeholder='Buscar en el detalle...'></div>"
        + f"<div class='people-wrapper'>{people_table(filtered)}</div></section></div>"
    )
    return layout(content, "MDO")


def detalle_page(query):
    rows = STATE["rows"]
    filters = {
        "ubicacion": "Ubicación",
        "centro": "Centro de Costo",
        "sexo": "Sexo",
        "contrato": "Modalidad de Contratación",
    }
    title_parts = []
    filtered = rows
    for param, field in filters.items():
        value = query.get(param, [""])[0]
        if value:
            filtered = [r for r in filtered if str(r.get(field)) == value]
            title_parts.append(f"{field}: {value}")
    title = "Detalle" if not title_parts else " | ".join(title_parts)
    content = (
        f"<section class='panel full'><div class='toolbar'><h2>{esc(title)} ({len(filtered)})</h2>"
        + "<input class='search' data-search placeholder='Buscar en el detalle...'></div>"
        + f"<div class='people-wrapper'>{people_table(filtered)}</div></section>"
    )
    return layout(content, "Detalle")


def antiguedad_page():
    rows = STATE["rows"]
    if not rows:
        return upload_page()
    groups = defaultdict(list)
    for r in rows:
        groups[r["Centro de Costo"]].append(r)
    table_rows = []
    for cc, people in sorted(groups.items()):
        ants = [p["Antiguedad"] for p in people if isinstance(p["Antiguedad"], (int, float))]
        table_rows.append(
            f"<tr><td><a href='/detalle?centro={urllib.parse.quote(str(cc))}'>{esc(cc)}</a></td>"
            f"<td class='num'>{len(people)}</td><td class='num'>{avg(ants)}</td>"
            f"<td class='num'>{sum(1 for a in ants if a < 1)}</td>"
            f"<td class='num'>{sum(1 for a in ants if 1 <= a < 5)}</td>"
            f"<td class='num'>{sum(1 for a in ants if 5 <= a < 10)}</td>"
            f"<td class='num'>{sum(1 for a in ants if a >= 10)}</td></tr>"
        )
    content = (
        "<section class='panel full'><h2>Antiguedad por centro de costo</h2>"
        "<table><thead><tr><th>Centro</th><th>Cantidad</th><th>Prom.</th><th>&lt;1</th><th>1 a 5</th><th>5 a 10</th><th>+10</th></tr></thead>"
        f"<tbody>{''.join(table_rows)}</tbody></table></section>"
    )
    return layout(content, "Antiguedad")


class Handler(BaseHTTPRequestHandler):
    def do_GET(self):
        parsed = urllib.parse.urlparse(self.path)
        query = urllib.parse.parse_qs(parsed.query)
        if parsed.path == "/":
            body = dashboard()
        elif parsed.path == "/general":
            body = general_page()
        elif parsed.path == "/mdo":
            body = mdo_page(query)
        elif parsed.path == "/detalle":
            body = detalle_page(query)
        elif parsed.path == "/antiguedad":
            body = antiguedad_page()
        else:
            self.send_error(404)
            return
        self.respond(body)

    def do_POST(self):
        if self.path != "/upload":
            self.send_error(404)
            return
        content_type = self.headers.get("Content-Type", "")
        length = int(self.headers.get("Content-Length", "0"))
        body = self.rfile.read(length)
        boundary = content_type.split("boundary=")[-1].encode()
        parts = body.split(b"--" + boundary)
        file_data = None
        filename = "input.xlsx"
        for part in parts:
            if b'name="file"' not in part:
                continue
            header, _, data = part.partition(b"\r\n\r\n")
            if b"filename=" in header:
                filename = header.split(b"filename=")[1].split(b"\r\n")[0].strip(b'"').decode("utf-8", "ignore")
            file_data = data.rsplit(b"\r\n", 1)[0]
        if not file_data:
            self.send_error(400, "No se recibio archivo")
            return
        parse_workbook(file_data, filename)
        self.send_response(303)
        self.send_header("Location", "/")
        self.end_headers()

    def respond(self, body):
        data = body.encode("utf-8")
        self.send_response(200)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def log_message(self, fmt, *args):
        return


def free_port(start=8765):
    for port in range(start, start + 50):
        with socket.socket() as s:
            try:
                s.bind(("127.0.0.1", port))
                return port
            except OSError:
                continue
    raise RuntimeError("No hay puertos libres")


if __name__ == "__main__":
    port = free_port()
    server = ThreadingHTTPServer(("127.0.0.1", port), Handler)
    print(f"App lista: http://127.0.0.1:{port}")
    print("Para cerrar, presiona Ctrl+C en esta ventana.")
    server.serve_forever()
